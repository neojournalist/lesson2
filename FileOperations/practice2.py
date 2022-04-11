from openpyxl import *
import statistics as stats

wb = load_workbook('/Users/altynai/Downloads/food.xlsx')
sheet = wb.worksheets[0]

cells = sheet['B12':'G28']

for row in cells:
    for data in row:
        print(data.value, end=" ")

cells2 = sheet.iter_rows(min_row=12, max_col=6, max_row=28)

for row in cells2:
    for data in row:
        print(data.value, end=" ")
    print(end="\n")
print('***********')

cells3 = sheet['F19':'G30']
listCells = []

for row in cells3:
    for data in row:
        listCells.append(data.value)

print(f'Median {stats.median(listCells)}')
print(f'Average {stats.mean(listCells)}')
print(f'Variance {stats.variance(listCells)}')
print(f'St Deviation {stats.stdev(listCells)}')

wb2 = Workbook()
sheet2 = wb2.active


dataInfo = (
    ('Region', 'City', 'Population in 1000', 'Mayor'),
    ('Naryn', 'Naryn', '300', 'ADB'),
    ('Osh', 'Osh', '700', 'KLP'),
    ('Talas', 'Talas', '100', 'GHN')
)

for row in dataInfo:
    sheet2.append(row)

wb2.save('KGZInfo.xlsx')






