import openpyxl
from openpyxl import *
from openpyxl.styles import *

wb = load_workbook('Book1.xlsx')
sheet = wb.worksheets[0]
name1Person = sheet['A2'].value
age1Person = sheet['B2'].value
print(f'First persons name: {sheet[A2].value} ')

#reading by cells

cells = sheet[A3:C3]
print(cells)

for info1, info2, info3 in cells:
    print(info1.value, info2.value, info3.value)

age = sheet ['B']
print(age)
columnAgeList = list()
for age in ages:
    if age.value == 'Vozrast':
        continue
    columnAgeList.append(age.value)

columnAgeList2 = [age.value for age in ages if age.value == 'Vozrast']

columnAgeList2 = []
for age in ages:
    if age.value =='Vozrast':
        continue
    columnAgeList2.append(age.value)

#creating an EXcel file

wb2 = Workbook()
sheet1 = wb2.worksheets[0]

nameStudentList = ['Nazgul', 'Kymbat', 'Nazima', 'Aiperi', 'Ayar']
countryList = ['Belgium', 'Poland', 'UK', 'USA', 'Ukraine']
infoStudentEnglishLevel = [80, 82, 75, 90, 67]

sheet1['A1'].font = Font(bold = True)
sheet1['B1'].font = Font(bold = True)
sheet1['c1'].font = Font(bold = True)



columnName1 = sheet1.cell(row=1, column=1).value = 'Student name'
columnName2 = sheet1.cell(row=1, column=2).value = 'Country'
columnName3 = sheet1.cell(row=1, column=3).value = 'English level'

counter = 2
for i in range(6):
    sheet1.cell(row = counter, column = 1).value = nameStudentList[i]
    sheet1.cell(row = counter, column = 2).value = countryList[i]
    sheet1.cell(row = counter, column = 3).value = infoStudentEnglishLevel[i]
    counter += 1

wb2.save('infoStudent.xlsx')
print('Successfully saved')




