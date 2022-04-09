from openpyxl import *
import statistics as stats

# отображение всех страниц с экзель
wb = load_workbook('/Users/altynai/Downloads/food.xlsx')
sheet = wb.worksheets[0]
allSheetNames = wb.sheetnames
print(allSheetNames)

# отображение вкладок по имени

sheetName = wb['Food_List']
sheetName2 = wb['math_values']

print(sheetName['B2'].value)

# Чтение большого промежутка
cells = sheetName['A2':'D17']

for row in cells:
    for data in row:
        print(data.value, end=" ")
    print(end="\n")

cells2 = sheetName.iter_rows(min_row=1, max_col=5, max_row=20)

for row in cells2:
    for data in row:
        print(data.value, end=" ")
    print(end="\n")
print('***********')

for row in sheetName.rows:
    for data in row:
        print(data.value, end=" ")
    print(end="\n")

allrows = sheetName.rows
print(next(allrows))

cells3 = sheetName['E5':'G5']

for row in cells3:
    for data in row:
        print(type(data.value))

allRows = sheet.rows
values = []

for row in allRows:
    for data in row:
        values.append(data.value)

print(values)

print(f'Number of values: {len(values)}')
print(f'Number of values: {max(values)}')
print(values[3:])
listNum = values[3:]
listNum.sort()
print(listNum)
print(f'Median {stats.median(values[3:])}')
print(f'Variance {stats.variance(values[3:])}')
print(f'St Dev {stats.stdev(values[3:])}')
print(f'Mean {stats.mean(values[3:])}')

"""
1)  Чтение промежутка 2 способами
- по координатам
- по указаниям строки и колонок
['B12':'G28']

2) ['F19':'G30'] - Прочитать в список
Получить макс значение
Найти
- Медиану
- Средн арифм значение
- Квадратное отклонение

3) Запись данных в файл экзель
Информацию про Кыргызстан
- Название областей
- Название крупных городов областей
- Население тыс(100,200)
- Мэр крупного города(наугад ФИО)
"""


